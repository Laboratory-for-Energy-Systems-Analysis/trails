from __future__ import annotations

import argparse
import re
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

CELL_REF = re.compile(r"(?<![A-Za-z0-9_])\$?([A-Z]{1,3})\$?([0-9]+)(?![A-Za-z0-9_])")


def _coord(match: re.Match[str]) -> str:
    return f"{match.group(1)}{match.group(2)}"


def _evaluate_workbook_formulas(src: Path, dst: Path) -> dict[str, int]:
    shutil.copy2(src, dst)
    workbook = load_workbook(dst, data_only=False)
    counts = {}

    for worksheet in workbook.worksheets:
        cache: dict[str, Any] = {}
        visiting: set[str] = set()

        def cell_value(coord: str) -> Any:
            coord = coord.replace("$", "")
            if coord in cache:
                return cache[coord]
            if coord in visiting:
                raise ValueError(f"Circular formula in {worksheet.title}!{coord}")

            visiting.add(coord)
            value = worksheet[coord].value
            if isinstance(value, str) and value.startswith("="):
                value = eval_formula(value[1:])
            visiting.remove(coord)
            cache[coord] = value
            return value

        def eval_formula(formula: str) -> Any:
            formula = formula.strip()
            match = CELL_REF.fullmatch(formula)
            if match:
                return cell_value(_coord(match))

            def replace_cell_ref(match: re.Match[str]) -> str:
                value = cell_value(_coord(match))
                if value is None:
                    value = 0
                return repr(value)

            expression = CELL_REF.sub(replace_cell_ref, formula)
            return eval(expression, {"__builtins__": {}}, {})

        formula_cells = [
            cell.coordinate
            for row in worksheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
        for coord in formula_cells:
            worksheet[coord].value = cell_value(coord)
        counts[worksheet.title] = len(formula_cells)

    workbook.save(dst)
    return counts


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("destination", type=Path)
    args = parser.parse_args()

    counts = _evaluate_workbook_formulas(
        args.source.expanduser().resolve(),
        args.destination.expanduser().resolve(),
    )
    print(args.destination.expanduser().resolve())
    print(counts)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
