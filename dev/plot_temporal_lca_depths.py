from __future__ import annotations

import argparse
import copy
import csv
import hashlib
import os
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from textwrap import shorten, wrap
from typing import Any

import numpy as np
import plotly.graph_objects as go
from datapackage import Package
from openpyxl import load_workbook
from plotly.subplots import make_subplots

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) in sys.path:
    sys.path.remove(str(REPO_ROOT))
sys.path.insert(0, str(REPO_ROOT))

from trails import Trails, get_lcia_method_names, plot_temporal_scores

DEFAULT_DATAPACKAGE = REPO_ROOT / "dev" / "trails_2026-05-18.zip"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "dev" / "temporal_lca_depth_plots_2026_05_18"
DEFAULT_METHOD = (
    "IPCC 2021 (incl. biogenic CO2) - climate change: total "
    "(incl. biogenic CO2) - global warming potential (GWP100)"
)
DEFAULT_ACTIVITY_INDICES = [
    26633,
    28292,
    24034,
    12342,
    4637,
    29249,
    12507,
    20515,
    14095,
    27773,
]
DEFAULT_INVENTORY_PATHS = sorted(REPO_ROOT.glob("dev/lci-*.xlsx"))
DEFAULT_DEPTHS = [1, 2, 3, 4, 5]
BASE_FONT_SIZE = 16
TITLE_FONT_SIZE = 22
SUBPLOT_TITLE_FONT_SIZE = 17
AXIS_TITLE_FONT_SIZE = 16
TICK_FONT_SIZE = 14
LEGEND_FONT_SIZE = 13
ROOT_COLOR_PALETTE = [
    "#1f77b4",
    "#ff7f0e",
    "#2ca02c",
    "#d62728",
    "#9467bd",
    "#8c564b",
    "#e377c2",
    "#7f7f7f",
    "#bcbd22",
    "#17becf",
    "#4e79a7",
    "#f28e2b",
    "#59a14f",
    "#e15759",
    "#b07aa1",
    "#9c755f",
    "#edc948",
    "#76b7b2",
]


@dataclass(frozen=True)
class ActivityDef:
    name: str
    reference_product: str
    location: str


def _slugify(value: str, max_length: int = 120) -> str:
    text = re.sub(r"[^A-Za-z0-9]+", "_", str(value)).strip("_").lower()
    return (text or "activity")[:max_length].strip("_")


def _ef_v31_methods(methods: list[str]) -> list[str]:
    return [
        method
        for method in methods
        if method.startswith("EF v3.1 -") and not method.startswith("EF v3.1 no LT")
    ]


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _read_activity_and_exchanges(ws) -> tuple[ActivityDef | None, list[dict[str, str]]]:
    activity_name = ""
    reference_product = ""
    location = ""

    for row in range(1, 80):
        key = _clean(ws.cell(row, 1).value).lower()
        value = _clean(ws.cell(row, 2).value)
        if key == "activity":
            activity_name = value
        elif key == "reference product":
            reference_product = value
        elif key == "location":
            location = value

    if not activity_name:
        return None, []

    header_row = None
    for row in range(1, 120):
        if (
            _clean(ws.cell(row, 1).value).lower() == "name"
            and _clean(ws.cell(row, 10).value).lower() == "type"
        ):
            header_row = row
            break

    exchanges: list[dict[str, str]] = []
    if header_row is not None:
        headers = [_clean(ws.cell(header_row, col).value) for col in range(1, 21)]
        for row in range(header_row + 1, 400):
            row_values = [_clean(ws.cell(row, col).value) for col in range(1, 21)]
            if not any(row_values):
                continue
            exchanges.append(
                {headers[i]: row_values[i] for i in range(len(headers)) if headers[i]}
            )

    return ActivityDef(activity_name, reference_product, location), exchanges


def _collect_terminal_activities(inventory_paths: list[Path]) -> list[ActivityDef]:
    all_activities: list[ActivityDef] = []
    activity_names: set[str] = set()
    technosphere_supplier_names: list[str] = []

    for path in inventory_paths:
        workbook = load_workbook(path, data_only=True)
        for worksheet in workbook.worksheets:
            activity, exchanges = _read_activity_and_exchanges(worksheet)
            if activity is None:
                continue
            all_activities.append(activity)
            activity_names.add(activity.name)
            for exchange in exchanges:
                if _clean(exchange.get("type", "")).lower() != "technosphere":
                    continue
                supplier_name = _clean(exchange.get("name", ""))
                if supplier_name:
                    technosphere_supplier_names.append(supplier_name)

    suppliers_used_by_imported = {
        supplier
        for supplier in technosphere_supplier_names
        if supplier in activity_names
    }
    terminal = [
        activity
        for activity in all_activities
        if activity.name not in suppliers_used_by_imported
    ]

    seen: set[tuple[str, str, str]] = set()
    out: list[ActivityDef] = []
    for activity in terminal:
        key = (activity.name, activity.reference_product, activity.location)
        if key in seen:
            continue
        seen.add(key)
        out.append(activity)
    return out


def _metadata_by_idx(trails: Trails) -> dict[int, dict]:
    if not trails.activity_indices:
        return {}
    first_label = next(iter(trails.activity_indices))
    return {
        int(key): value for key, value in trails.activity_indices[first_label].items()
    }


def _match_activity_indices(
    trails: Trails,
    targets: list[ActivityDef],
) -> dict[ActivityDef, int]:
    by_idx = _metadata_by_idx(trails)
    out: dict[ActivityDef, int] = {}

    for target in targets:
        exact: list[int] = []
        for index, metadata in by_idx.items():
            if not isinstance(metadata, dict):
                continue
            if _clean(metadata.get("name")) != target.name:
                continue
            if _clean(metadata.get("reference product")) != target.reference_product:
                continue
            if _clean(metadata.get("location")) != target.location:
                continue
            exact.append(index)
        if len(exact) == 1:
            out[target] = exact[0]
            continue

        by_name = [
            index
            for index, metadata in by_idx.items()
            if isinstance(metadata, dict)
            and _clean(metadata.get("name")) == target.name
        ]
        if len(by_name) == 1:
            out[target] = by_name[0]

    return out


def _wrap_title(value: str, width: int) -> str:
    lines = wrap(
        str(value), width=width, break_long_words=False, break_on_hyphens=False
    )
    return "<br>".join(lines) if lines else str(value)


def _method_title_label(method: str) -> str:
    if "IPCC 2021" in method and "GWP100" in method:
        if "incl. biogenic CO2" in method:
            return "IPCC 2021 GWP100 (incl. biogenic CO2)"
        if "excl. biogenic CO2" in method:
            return "IPCC 2021 GWP100 (excl. biogenic CO2)"
        return "IPCC 2021 GWP100"
    return shorten(method, width=110, placeholder="...")


def _stable_color(key: str) -> str:
    digest = hashlib.blake2b(str(key).encode("utf-8"), digest_size=2).digest()
    idx = int.from_bytes(digest, "big") % len(ROOT_COLOR_PALETTE)
    return ROOT_COLOR_PALETTE[idx]


def _hex_to_rgba(color: str, alpha: float) -> str:
    color = str(color).lstrip("#")
    if len(color) != 6:
        return f"rgba(31, 119, 180, {alpha:.3f})"
    r = int(color[0:2], 16)
    g = int(color[2:4], 16)
    b = int(color[4:6], 16)
    return f"rgba({r}, {g}, {b}, {alpha:.3f})"


def _style_trace_consistently(trace: go.BaseTraceType) -> None:
    name = str(getattr(trace, "name", "") or "")
    legendgroup = str(getattr(trace, "legendgroup", "") or name)
    if name == "Cumulative total":
        color = "#111827"
    elif name == "Static score":
        color = "#dc2626"
    elif name == "Total annual impact":
        color = "#2563eb"
    else:
        color = _stable_color(legendgroup)

    if hasattr(trace, "line"):
        trace.line.color = color
    if hasattr(trace, "marker"):
        trace.marker.color = color
    if getattr(trace, "fill", None):
        trace.fillcolor = _hex_to_rgba(color, 0.34)


def _activity_metadata(
    trails: Trails, activity_index: int, reference_year: int
) -> dict:
    labels: list[str] = []
    try:
        labels.append(str(trails._map_year_to_scenario_year(int(reference_year))))
    except Exception:
        pass
    labels.extend(str(label) for label in getattr(trails, "scenario_labels", []))

    seen: set[str] = set()
    for label in labels:
        if label in seen:
            continue
        seen.add(label)
        mapping = trails.activity_indices.get(label, {})
        meta = mapping.get(int(activity_index))
        if isinstance(meta, dict):
            return dict(meta)

    for mapping in trails.activity_indices.values():
        meta = mapping.get(int(activity_index))
        if isinstance(meta, dict):
            return dict(meta)

    return {}


def _activity_label(
    trails: Trails,
    activity_index: int,
    reference_year: int,
    *,
    max_length: int | None = None,
) -> str:
    meta = _activity_metadata(trails, activity_index, reference_year)
    name = _clean(meta.get("name")) or f"Activity {activity_index}"
    reference_product = _clean(meta.get("reference product"))
    location = _clean(meta.get("location"))
    parts = [name]
    if reference_product:
        parts.append(reference_product)
    if location:
        parts.append(location)
    label = " | ".join(parts)
    if max_length is not None:
        return shorten(label, width=max_length, placeholder="...")
    return label


def _validate_activity_index(trails: Trails, activity_index: int) -> None:
    if trails.A is None:
        raise RuntimeError("Trails.A is not initialized.")
    n_activities = int(trails.A.shape[1])
    if int(activity_index) < 0 or int(activity_index) >= n_activities:
        raise ValueError(
            f"Activity index {activity_index} is outside the A matrix activity "
            f"axis with length {n_activities}."
        )


def _score_to_float(score: object) -> float:
    arr = np.asarray(score, dtype=float)
    if arr.size == 0:
        raise ValueError("Static score is empty.")
    return float(arr.ravel()[0])


def _score_to_method_map(score: object, methods: list[str]) -> dict[str, float]:
    arr = np.asarray(score, dtype=float).ravel()
    if arr.size == 0:
        raise ValueError("Static score is empty.")
    if arr.size == 1 and len(methods) == 1:
        return {methods[0]: float(arr[0])}
    if arr.size != len(methods):
        raise ValueError(
            "Static score length does not match method count: "
            f"{arr.size} scores for {len(methods)} methods."
        )
    return {method: float(arr[idx]) for idx, method in enumerate(methods)}


def _scores_reduced_to_method_root_year(trails: Trails):
    if trails.scores is None:
        raise RuntimeError("trails.scores is None after temporal LCA.")

    data = trails.scores
    keep_dims = [dim for dim in ("method", "root activity", "year") if dim in data.dims]
    reduce_dims = [dim for dim in data.dims if dim not in keep_dims]
    if reduce_dims:
        data = data.sum(dim=reduce_dims)
    if data.dims != tuple(keep_dims):
        data = data.transpose(*keep_dims)
    return data


def _select_method_by_position(data, method: str):
    if "method" not in data.dims:
        return data
    methods = [str(value) for value in data.coords["method"].values.tolist()]
    try:
        method_idx = methods.index(str(method))
    except ValueError as exc:
        raise ValueError(
            f"Method '{method}' not found in scores. Available: {methods}"
        ) from exc
    return data.isel(method=method_idx, drop=True)


def _temporal_score_by_year_from_scores(
    scores,
    method: str,
) -> tuple[np.ndarray, np.ndarray]:
    data = _select_method_by_position(scores, method)

    reduce_dims = [dim for dim in data.dims if dim != "year"]
    if reduce_dims:
        data = data.sum(dim=reduce_dims)
    if data.dims != ("year",):
        data = data.transpose("year")

    years = np.asarray(data.coords["year"].values, dtype=int)
    values_raw = data.data
    if hasattr(values_raw, "todense"):
        values = np.asarray(values_raw.todense(), dtype=float).ravel()
    else:
        values = np.asarray(data.values, dtype=float).ravel()
    return years, values


def _temporal_score_by_year(
    trails: Trails, method: str
) -> tuple[np.ndarray, np.ndarray]:
    return _temporal_score_by_year_from_scores(
        _scores_reduced_to_method_root_year(trails),
        method,
    )


def _pruned_scores_for_method(scores, method: str):
    scores = _select_method_by_position(scores, method)

    if "root activity" not in scores.dims:
        return scores

    extra_dims = [dim for dim in scores.dims if dim not in ("root activity", "year")]
    if extra_dims:
        scores = scores.sum(dim=extra_dims)

    if scores.dims != ("root activity", "year"):
        scores = scores.transpose("root activity", "year")

    values = scores.data
    active_positions: np.ndarray | None = None
    if hasattr(values, "coords") and hasattr(values, "nnz"):
        if int(values.nnz) == 0:
            active_positions = np.array([], dtype=int)
        else:
            active_positions = np.unique(np.asarray(values.coords[0], dtype=int))
    else:
        dense = np.asarray(scores.values, dtype=float)
        if dense.ndim == 2:
            active_positions = np.flatnonzero(np.any(dense != 0.0, axis=1))

    if active_positions is None:
        return scores
    if active_positions.size == scores.sizes["root activity"]:
        return scores
    return scores.isel({"root activity": active_positions})


def _plot_temporal_scores_for_method(
    trails: Trails,
    *,
    method: str,
    scores,
    **kwargs,
):
    pruned_scores = _pruned_scores_for_method(scores, method)
    previous_scores = trails.scores
    previous_characterized_inventory = trails.characterized_inventory
    try:
        trails.characterized_inventory = None
        trails.scores = pruned_scores
        return plot_temporal_scores(trails=trails, method=None, **kwargs)
    finally:
        trails.scores = previous_scores
        trails.characterized_inventory = previous_characterized_inventory


def _trim_years(
    years: np.ndarray,
    values: np.ndarray,
    *,
    year_start: int | None,
    year_end: int | None,
) -> tuple[np.ndarray, np.ndarray]:
    mask = np.ones(years.shape, dtype=bool)
    if year_start is not None:
        mask &= years >= int(year_start)
    if year_end is not None:
        mask &= years <= int(year_end)
    return years[mask], values[mask]


def _observed_year_bounds(years: np.ndarray, values: np.ndarray) -> tuple[int, int]:
    if years.size == 0:
        raise ValueError("Cannot determine observed year bounds from empty scores.")
    max_abs = float(np.nanmax(np.abs(values))) if values.size else 0.0
    tol = np.finfo(float).eps * max(1.0, max_abs) * 10.0
    active = np.abs(values) > tol
    if not np.any(active):
        y0 = int(years[0])
        return y0, y0
    active_years = years[active]
    return int(active_years[0]), int(active_years[-1])


def _apply_axis_ranges_from_source(
    target: go.Figure,
    source: go.Figure,
    *,
    row: int,
) -> None:
    primary_range = source.layout.yaxis.range
    secondary_range = source.layout.yaxis2.range
    if primary_range is not None:
        target.update_yaxes(
            range=list(primary_range), row=row, col=1, secondary_y=False
        )
    if secondary_range is not None:
        target.update_yaxes(
            range=list(secondary_range),
            row=row,
            col=1,
            secondary_y=True,
        )


def _add_total_annual_trace(
    fig: go.Figure,
    years: np.ndarray,
    annual: np.ndarray,
) -> None:
    if years.size == 0:
        return
    fig.add_trace(
        go.Scatter(
            x=[int(year) for year in years],
            y=[float(value) for value in annual],
            name="Total annual impact",
            legendgroup="annual-total",
            showlegend=True,
            mode="lines",
            line=dict(color="#2563eb", width=2.5, dash="dot"),
            yaxis="y2",
            hovertemplate=(
                "<b>Total annual impact</b><br>"
                "Year: %{x}<br>"
                "Annual impact: %{y:.6g}<extra></extra>"
            ),
        )
    )


def _extend_secondary_trace_to_window(
    trace,
    *,
    x_start: int,
    x_end: int,
) -> None:
    if getattr(trace, "yaxis", None) != "y2":
        return

    name = str(getattr(trace, "name", "") or "")
    if name == "Static score":
        raw_y = getattr(trace, "y", None)
        y = [] if raw_y is None else list(raw_y)
        if not y:
            return
        trace.x = [int(x_start), int(x_end)]
        trace.y = [float(y[0]), float(y[0])]
        return

    if name != "Cumulative total":
        return

    raw_x = getattr(trace, "x", None)
    raw_y = getattr(trace, "y", None)
    x = [] if raw_x is None else [int(value) for value in raw_x]
    y = [] if raw_y is None else [float(value) for value in raw_y]
    if not x or not y:
        return

    if int(x[0]) > int(x_start):
        x.insert(0, int(x_start))
        y.insert(0, 0.0)
    if int(x[-1]) < int(x_end):
        x.append(int(x_end))
        y.append(float(y[-1]))
    trace.x = x
    trace.y = y


def _fallback_depth_figure(
    *,
    years: np.ndarray,
    annual: np.ndarray,
    static_score: float,
    reference_year: int,
) -> go.Figure:
    if years.size == 0:
        years = np.array([int(reference_year)], dtype=int)
        annual = np.array([0.0], dtype=float)

    cumulative = np.cumsum(np.asarray(annual, dtype=float))
    fig = go.Figure()
    fig.add_trace(
        go.Scatter(
            x=[int(year) for year in years],
            y=[0.0 for _ in years],
            name="No root-attributed annual impacts",
            showlegend=True,
            mode="lines",
            line=dict(color="#9ca3af", width=1),
            hoverinfo="skip",
        )
    )
    fig.add_trace(
        go.Scatter(
            x=[int(year) for year in years],
            y=[float(value) for value in cumulative],
            name="Cumulative total",
            legendgroup="cumulative-total",
            showlegend=True,
            mode="lines",
            line=dict(color="#111827", width=2),
            yaxis="y2",
            hovertemplate=(
                "<b>Cumulative total</b><br>"
                "Year: %{x}<br>"
                "Cumulative impact: %{y:.6g}<extra></extra>"
            ),
        )
    )
    fig.add_trace(
        go.Scatter(
            x=[int(years[0]), int(years[-1])],
            y=[float(static_score), float(static_score)],
            name="Static score",
            legendgroup="static-score",
            showlegend=True,
            mode="lines",
            line=dict(color="#dc2626", width=2, dash="dash"),
            yaxis="y2",
            hovertemplate=(
                "<b>Static score</b><br>" "Static score: %{y:.6g}<extra></extra>"
            ),
        )
    )
    fig.update_layout(
        yaxis=dict(range=[-1.0, 1.0]),
        yaxis2=dict(
            range=[min(0.0, float(static_score)), max(1.0, float(static_score))]
        ),
    )
    return fig


def _routing_graph_summary(trails: Trails) -> str:
    graph = getattr(trails, "graph", None)
    if graph is None:
        return "nodes=0 edges=0"

    max_depth = 0
    frontier_nodes = 0
    direct_bio_nodes = 0
    for _, data in graph.nodes(data=True):
        max_depth = max(max_depth, int(data.get("depth", 0)))
        if data.get("frontier_amount"):
            frontier_nodes += 1
        if data.get("direct_bio_amount"):
            direct_bio_nodes += 1

    return (
        f"nodes={int(graph.number_of_nodes()):,} "
        f"edges={int(graph.number_of_edges()):,} "
        f"max_depth={max_depth} "
        f"frontier_nodes={frontier_nodes:,} "
        f"direct_bio_nodes={direct_bio_nodes:,}"
    )


def _run_temporal_lca_for_depth(
    trails: Trails,
    *,
    activity_index: int,
    reference_year: int,
    amount: float,
    depth: int,
    methods: list[str],
    show_progress: bool,
    solver_mode: str,
    iterative_rtol: float,
    iterative_maxiter: int | None,
    iterative_restart: int | None,
    ei_version: str,
    fallback_solver_mode: str | None,
    routing_min_amount: float,
) -> None:
    print(
        f"    routing start: depth={depth}, min_amount={routing_min_amount:g}",
        flush=True,
    )
    routing_t0 = time.perf_counter()
    trails.temporal_routing(
        start_year=int(reference_year),
        start_act_idx=int(activity_index),
        amount=float(amount),
        max_depth=int(depth),
        min_amount=float(routing_min_amount),
        show_progress=bool(show_progress),
        attribute_to_roots=True,
    )
    print(
        "    routing done in "
        f"{time.perf_counter() - routing_t0:.1f}s "
        f"({_routing_graph_summary(trails)})",
        flush=True,
    )
    print(
        f"    lca start: {len(methods)} method(s), solver_mode={solver_mode}",
        flush=True,
    )
    lca_t0 = time.perf_counter()
    try:
        trails.lca(
            methods=methods,
            show_progress=bool(show_progress),
            attribute_to_roots=True,
            compute_score=True,
            store_inventory=False,
            solver_mode=solver_mode,
            iterative_rtol=float(iterative_rtol),
            iterative_maxiter=iterative_maxiter,
            iterative_restart=iterative_restart,
            ei_version=ei_version,
        )
        print(f"    lca done in {time.perf_counter() - lca_t0:.1f}s", flush=True)
    except RuntimeError as exc:
        if (
            solver_mode != "iterative"
            or fallback_solver_mode is None
            or fallback_solver_mode == solver_mode
            or "GMRES failed to converge" not in str(exc)
        ):
            raise
        print(
            "  iterative solver did not converge; retrying depth="
            f"{depth} with solver_mode={fallback_solver_mode}",
            flush=True,
        )
        fallback_t0 = time.perf_counter()
        trails.lca(
            methods=methods,
            show_progress=bool(show_progress),
            attribute_to_roots=True,
            compute_score=True,
            store_inventory=False,
            solver_mode=fallback_solver_mode,
            ei_version=ei_version,
        )
        print(
            "    fallback lca done in " f"{time.perf_counter() - fallback_t0:.1f}s",
            flush=True,
        )


def _write_summary_csv(rows: list[dict[str, object]], csv_path: Path) -> None:
    if not rows:
        return
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def _plot_activity_depths(
    *,
    trails: Trails,
    activity_index: int,
    activity_label: str,
    method: str,
    static_score: float,
    plotly_figures_by_depth: dict[int, go.Figure],
    observed_bounds_by_depth: dict[int, tuple[int, int]],
    reference_year: int,
    year_start: int | None,
    year_end: int | None,
    output_dir: Path,
    width: int,
    height: int,
) -> Path:
    depths = sorted(plotly_figures_by_depth)
    observed_start = min(bounds[0] for bounds in observed_bounds_by_depth.values())
    observed_end = max(bounds[1] for bounds in observed_bounds_by_depth.values())
    x_start = int(year_start) if year_start is not None else int(observed_start)
    x_end = int(year_end) if year_end is not None else int(observed_end)
    if x_start == x_end:
        x_start -= 1
        x_end += 1

    subplot_titles = [f"Depth {depth}" for depth in depths]
    fig = make_subplots(
        rows=len(depths),
        cols=1,
        shared_xaxes=True,
        specs=[[{"secondary_y": True}] for _ in depths],
        vertical_spacing=0.055,
        subplot_titles=subplot_titles,
    )

    seen_legend_entries: set[str] = set()
    for row, depth in enumerate(depths, start=1):
        depth_fig = plotly_figures_by_depth[depth]
        for source_trace in depth_fig.data:
            trace = copy.deepcopy(source_trace)
            _extend_secondary_trace_to_window(trace, x_start=x_start, x_end=x_end)
            _style_trace_consistently(trace)
            secondary_y = getattr(trace, "yaxis", None) == "y2"
            legend_key = str(trace.legendgroup or trace.name or trace.uid)
            source_showlegend = getattr(source_trace, "showlegend", None)
            wants_legend = (
                True if source_showlegend is None else bool(source_showlegend)
            )
            trace.showlegend = wants_legend and legend_key not in seen_legend_entries
            if trace.showlegend:
                seen_legend_entries.add(legend_key)
            fig.add_trace(trace, row=row, col=1, secondary_y=secondary_y)

        _apply_axis_ranges_from_source(fig, depth_fig, row=row)
        fig.add_vline(
            x=int(reference_year),
            line_width=1.5,
            line_dash="dash",
            line_color="#6b7280",
            row=row,
            col=1,
        )
        fig.update_yaxes(
            title_text="Annual impact",
            title_font=dict(size=AXIS_TITLE_FONT_SIZE),
            tickfont=dict(size=TICK_FONT_SIZE),
            row=row,
            col=1,
            secondary_y=False,
        )
        fig.update_yaxes(
            title_text="Cumulative / static",
            title_font=dict(size=AXIS_TITLE_FONT_SIZE),
            tickfont=dict(size=TICK_FONT_SIZE),
            row=row,
            col=1,
            secondary_y=True,
        )

    title_label = _wrap_title(shorten(activity_label, width=150, placeholder="..."), 82)
    method_label = _wrap_title(_method_title_label(method), 96)
    for annotation in fig.layout.annotations:
        annotation.font = dict(size=SUBPLOT_TITLE_FONT_SIZE)
    fig.update_layout(
        width=int(width),
        height=int(height),
        template="plotly_white",
        hovermode="x unified",
        font=dict(size=BASE_FONT_SIZE),
        title=dict(
            text=(
                f"Activity {activity_index}<br>{title_label}<br>"
                f"<sup>{method_label}</sup>"
            ),
            x=0.5,
            xanchor="center",
            font=dict(size=TITLE_FONT_SIZE),
        ),
        legend=dict(
            orientation="h",
            x=0.5,
            xanchor="center",
            y=-0.08,
            yanchor="top",
            font=dict(size=LEGEND_FONT_SIZE),
            entrywidth=230,
            entrywidthmode="pixels",
        ),
        margin=dict(l=78, r=92, t=195, b=140),
    )
    fig.update_xaxes(
        title_text="Year",
        title_font=dict(size=AXIS_TITLE_FONT_SIZE),
        tickfont=dict(size=TICK_FONT_SIZE),
        row=len(depths),
        col=1,
    )
    fig.update_xaxes(range=[x_start, x_end])

    safe_name = _slugify(activity_label)
    safe_method = _slugify(method, max_length=140)
    depth_label = "_".join(str(depth) for depth in depths)
    activity_dir = output_dir / f"idx_{int(activity_index)}_{safe_name}"
    activity_dir.mkdir(parents=True, exist_ok=True)
    path = activity_dir / f"{safe_method}_depths_{depth_label}.png"
    try:
        fig.write_image(str(path))
    except ValueError as exc:
        if "kaleido" in str(exc).lower():
            raise RuntimeError(
                "Plotly PNG export requires kaleido in the active environment. "
                "Install it in the trails conda environment if needed."
            ) from exc
        raise
    return path


def run(args: argparse.Namespace) -> int:
    if args.lcia_json is not None:
        lcia_json = Path(args.lcia_json).expanduser().resolve()
        if not lcia_json.exists():
            raise FileNotFoundError(f"LCIA JSON not found: {lcia_json}")
        os.environ["TRAILS_LCIA_EI312_JSON"] = str(lcia_json)

    available_methods = get_lcia_method_names(ei_version=str(args.ei_version))
    requested_methods = (
        [str(method) for method in args.methods] if args.methods else [str(args.method)]
    )
    if bool(args.include_ef_v31):
        requested_methods.extend(_ef_v31_methods(available_methods))
    requested_methods = list(dict.fromkeys(requested_methods))

    missing_methods = [
        method for method in requested_methods if method not in available_methods
    ]
    if missing_methods:
        matches = [name for name in available_methods if "IPCC 2021" in name][:12]
        raise ValueError(
            "LCIA method(s) not found for ecoinvent "
            f"{args.ei_version}:\n- " + "\n- ".join(missing_methods) + "\n"
            f"Nearby IPCC 2021 methods:\n- " + "\n- ".join(matches)
        )
    if not requested_methods:
        raise ValueError("At least one LCIA method is required.")
    print(f"LCIA methods: {len(requested_methods)}")
    for method in requested_methods:
        print(f"  {method}")

    print(f"Loading datapackage: {args.datapackage}")
    trails = Trails(
        Package(str(args.datapackage)),
        interpolate_annual=True,
        cache_interpolation=not bool(args.no_cache_interpolation),
        interpolation_start_year_offset=int(args.interpolation_start_year_offset),
        interpolation_end_year_offset=int(args.interpolation_end_year_offset),
    )

    activity_indices = list(args.activity_indices)
    selected_inventories = list(args.inventories)
    if bool(args.all_dev_lci_inventories):
        selected_inventories.extend(DEFAULT_INVENTORY_PATHS)
    inventory_paths = [
        Path(path).expanduser().resolve() for path in selected_inventories
    ]
    inventory_paths = list(dict.fromkeys(inventory_paths))
    if inventory_paths:
        missing = [path for path in inventory_paths if not path.exists()]
        if missing:
            raise FileNotFoundError(
                "Inventory file(s) not found:\n- "
                + "\n- ".join(str(path) for path in missing)
            )
        print("Importing Excel inventories:")
        for path in inventory_paths:
            print(f"  {path}")
        trails.import_excel_inventory([str(path) for path in inventory_paths])

    if bool(args.include_terminal_inventory_activities):
        if not inventory_paths:
            raise ValueError(
                "--include-terminal-inventory-activities requires --inventories."
            )
        terminal_activities = _collect_terminal_activities(inventory_paths)
        terminal_index_map = _match_activity_indices(trails, terminal_activities)
        missing_terminals = [
            activity
            for activity in terminal_activities
            if activity not in terminal_index_map
        ]
        if missing_terminals:
            print("Warning: could not match terminal imported activities:")
            for activity in missing_terminals:
                print(
                    "  "
                    f"{activity.name} | {activity.reference_product} | "
                    f"{activity.location}"
                )
        print("Terminal imported activities:")
        for activity, index in sorted(
            terminal_index_map.items(), key=lambda item: item[1]
        ):
            print(
                f"  {index}: {activity.name} | "
                f"{activity.reference_product} | {activity.location}"
            )
            activity_indices.append(int(index))

    activity_indices = list(dict.fromkeys(int(index) for index in activity_indices))

    depths = sorted({int(depth) for depth in args.depths})
    if not depths:
        raise ValueError("At least one depth is required.")
    if any(depth < 0 for depth in depths):
        raise ValueError("Depths must be non-negative integers.")

    if int(args.plot_window_years) < 0:
        raise ValueError("--plot-window-years must be non-negative.")
    plot_year_start = (
        int(args.year_start)
        if args.year_start is not None
        else int(args.reference_year) - int(args.plot_window_years)
    )
    plot_year_end = (
        int(args.year_end)
        if args.year_end is not None
        else int(args.reference_year) + int(args.plot_window_years)
    )
    if plot_year_start > plot_year_end:
        raise ValueError("Plot year start must be <= plot year end.")

    output_dir = Path(args.output_dir).expanduser().resolve()
    csv_path = Path(args.results_csv).expanduser().resolve()
    rows: list[dict[str, object]] = []
    total = len(activity_indices)
    for pos, activity_index in enumerate(activity_indices, start=1):
        activity_index = int(activity_index)
        _validate_activity_index(trails, activity_index)
        label = _activity_label(trails, activity_index, int(args.reference_year))
        print(f"\n[{pos}/{total}] Activity {activity_index}: {label}", flush=True)

        t0 = time.perf_counter()
        trails.static_lca(
            year=int(args.reference_year),
            act_idx=activity_index,
            methods=requested_methods,
            amount=float(args.amount),
            ei_version=str(args.ei_version),
        )
        static_scores = _score_to_method_map(trails.static_score, requested_methods)
        print(f"  static_lca: {len(static_scores)} method score(s)", flush=True)

        plotly_figures_by_method_depth: dict[str, dict[int, go.Figure]] = {
            method: {} for method in requested_methods
        }
        observed_bounds_by_method_depth: dict[str, dict[int, tuple[int, int]]] = {
            method: {} for method in requested_methods
        }
        temporal_totals: dict[str, dict[int, float]] = {
            method: {} for method in requested_methods
        }
        window_temporal_totals: dict[str, dict[int, float]] = {
            method: {} for method in requested_methods
        }
        for depth in depths:
            print(f"  depth={depth}: temporal_routing + lca", flush=True)
            _run_temporal_lca_for_depth(
                trails,
                activity_index=activity_index,
                reference_year=int(args.reference_year),
                amount=float(args.amount),
                depth=int(depth),
                methods=requested_methods,
                show_progress=bool(args.show_progress),
                solver_mode=str(args.solver_mode),
                iterative_rtol=float(args.iterative_rtol),
                iterative_maxiter=args.iterative_maxiter,
                iterative_restart=args.iterative_restart,
                ei_version=str(args.ei_version),
                fallback_solver_mode=(
                    None
                    if str(args.fallback_solver_mode) == "none"
                    else str(args.fallback_solver_mode)
                ),
                routing_min_amount=float(args.routing_min_amount),
            )
            plot_year_range = (
                int(plot_year_start),
                int(plot_year_end),
            )
            print("    reducing scores for plotting", flush=True)
            scores_for_plotting = _scores_reduced_to_method_root_year(trails)
            for method_pos, method in enumerate(requested_methods, start=1):
                years, annual = _temporal_score_by_year_from_scores(
                    scores_for_plotting,
                    method,
                )
                print(
                    f"    plotting method {method_pos}/{len(requested_methods)}",
                    flush=True,
                )
                try:
                    depth_fig = _plot_temporal_scores_for_method(
                        trails=trails,
                        method=method,
                        scores=scores_for_plotting,
                        stacked=bool(args.stacked),
                        legend_top_n=int(args.legend_top_n),
                        show_flow_contributions=bool(args.show_flow_contributions),
                        title="",
                        method_label="Annual impact",
                        cumulative=False,
                        width=int(args.width),
                        height=max(320, int(args.height / max(1, len(depths)))),
                        year_tick=int(args.year_tick),
                        # Keep the full time series here so the cumulative trace
                        # starts from the first observed impact. The combined figure
                        # clips the x-axis to plot_year_range after traces are built.
                        year_range=None,
                        reference_year=int(args.reference_year),
                        show_cumulative_axis=True,
                        cumulative_axis_label="Cumulative impact",
                        static_score=float(static_scores[method]),
                        static_score_label="Static score",
                        static_score_dash="dash",
                        static_score_color="#dc2626",
                    )
                    if isinstance(depth_fig, list):
                        if not depth_fig:
                            raise RuntimeError(
                                "plot_temporal_scores returned no figures."
                            )
                        depth_fig = depth_fig[0]
                except ValueError as exc:
                    if "No scores_by_first_level_child found" not in str(exc):
                        raise
                    depth_fig = _fallback_depth_figure(
                        years=years,
                        annual=annual,
                        static_score=float(static_scores[method]),
                        reference_year=int(args.reference_year),
                    )
                if bool(args.show_total_annual_line):
                    _add_total_annual_trace(depth_fig, years, annual)
                plotly_figures_by_method_depth[method][int(depth)] = depth_fig

                plot_years, plot_annual = _trim_years(
                    years,
                    annual,
                    year_start=plot_year_start,
                    year_end=plot_year_end,
                )
                observed_bounds_by_method_depth[method][int(depth)] = (
                    _observed_year_bounds(plot_years, plot_annual)
                )
                full_temporal_total = float(np.sum(annual))
                window_temporal_total = float(np.sum(plot_annual))
                temporal_totals[method][int(depth)] = full_temporal_total
                window_temporal_totals[method][int(depth)] = window_temporal_total

            del scores_for_plotting
            print(f"  depth={depth}: plotted {len(requested_methods)} method(s)")

        plot_paths_by_method: dict[str, Path] = {}
        for method in requested_methods:
            path = _plot_activity_depths(
                trails=trails,
                activity_index=activity_index,
                activity_label=label,
                method=method,
                static_score=static_scores[method],
                plotly_figures_by_depth=plotly_figures_by_method_depth[method],
                observed_bounds_by_depth=observed_bounds_by_method_depth[method],
                reference_year=int(args.reference_year),
                year_start=plot_year_start,
                year_end=plot_year_end,
                output_dir=output_dir,
                width=int(args.width),
                height=int(args.height),
            )
            plot_paths_by_method[method] = path
        activity_dir = next(iter(plot_paths_by_method.values())).parent
        print(
            f"  wrote {len(plot_paths_by_method)} indicator figure(s) to "
            f"{activity_dir}"
        )
        print(f"  activity done in {time.perf_counter() - t0:.1f}s")

        for method in requested_methods:
            row: dict[str, object] = {
                "activity_index": activity_index,
                "activity_label": label,
                "method": method,
                "reference_year": int(args.reference_year),
                "amount": float(args.amount),
                "static_score": static_scores[method],
                "plot_path": str(plot_paths_by_method[method]),
            }
            for depth in depths:
                row[f"temporal_depth_{depth}"] = temporal_totals[method][int(depth)]
                row[f"delta_depth_{depth}_vs_static"] = (
                    temporal_totals[method][int(depth)] - static_scores[method]
                )
                row[f"window_temporal_depth_{depth}"] = window_temporal_totals[method][
                    int(depth)
                ]
            rows.append(row)
        _write_summary_csv(rows, csv_path)

    print(f"\nWrote summary CSV: {csv_path}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Run static and temporal Trails LCAs for selected activity indices and "
            "write one PNG per activity with one subplot per temporal depth."
        )
    )
    parser.add_argument("--datapackage", type=Path, default=DEFAULT_DATAPACKAGE)
    parser.add_argument(
        "--activity-indices",
        nargs="+",
        type=int,
        default=DEFAULT_ACTIVITY_INDICES,
    )
    parser.add_argument(
        "--inventories",
        nargs="*",
        type=Path,
        default=[],
        help="Excel inventory files to import before matching activity indices.",
    )
    parser.add_argument(
        "--all-dev-lci-inventories",
        action="store_true",
        help="Use all dev/lci-*.xlsx files as --inventories.",
    )
    parser.add_argument(
        "--include-terminal-inventory-activities",
        action="store_true",
        help=(
            "Append terminal activities from imported Excel inventories. A terminal "
            "activity is one not used as a technosphere supplier by another "
            "imported activity in the same workbook set."
        ),
    )
    parser.add_argument("--depths", nargs="+", type=int, default=DEFAULT_DEPTHS)
    parser.add_argument("--method", type=str, default=DEFAULT_METHOD)
    parser.add_argument(
        "--methods",
        nargs="+",
        type=str,
        default=None,
        help=(
            "Explicit LCIA methods to use. If omitted, --method is used. "
            "Use --include-ef-v31 to append all regular EF v3.1 methods."
        ),
    )
    parser.add_argument(
        "--include-ef-v31",
        action="store_true",
        help="Append all methods starting with 'EF v3.1 -' and exclude 'EF v3.1 no LT'.",
    )
    parser.add_argument("--ei-version", type=str, default="3.12")
    parser.add_argument(
        "--lcia-json",
        type=Path,
        default=None,
        help=(
            "Optional path to an ecoinvent 3.12-compatible LCIA JSON. By default "
            "Trails uses trails/data/lcia_ei312.json."
        ),
    )
    parser.add_argument("--reference-year", type=int, default=2035)
    parser.add_argument("--amount", type=float, default=1.0)
    parser.add_argument(
        "--year-start",
        type=int,
        default=None,
        help=(
            "Optional lower x-axis bound. Defaults to reference year minus "
            "--plot-window-years."
        ),
    )
    parser.add_argument(
        "--year-end",
        type=int,
        default=None,
        help=(
            "Optional upper x-axis bound. Defaults to reference year plus "
            "--plot-window-years."
        ),
    )
    parser.add_argument(
        "--plot-window-years",
        type=int,
        default=40,
        help=(
            "Half-width of the default plot window around the reference year. "
            "Ignored for bounds explicitly set with --year-start/--year-end."
        ),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
    )
    parser.add_argument(
        "--results-csv",
        type=Path,
        default=DEFAULT_OUTPUT_DIR / "temporal_lca_depth_scores.csv",
    )
    parser.add_argument("--width", type=int, default=1100)
    parser.add_argument("--height", type=int, default=1700)
    parser.add_argument("--legend-top-n", type=int, default=7)
    parser.add_argument("--year-tick", type=int, default=5)
    parser.add_argument(
        "--stacked",
        action=argparse.BooleanOptionalAction,
        default=False,
        help="Use stacked traces in plot_temporal_scores().",
    )
    parser.add_argument(
        "--show-flow-contributions",
        action="store_true",
        help="Pass show_flow_contributions=True to plot_temporal_scores().",
    )
    parser.add_argument(
        "--show-total-annual-line",
        action=argparse.BooleanOptionalAction,
        default=False,
        help=(
            "Overlay the total annual impact on the secondary y-axis. This helps "
            "show small annual net uptake values that are hidden by large "
            "root-attributed annual pulses."
        ),
    )
    parser.add_argument("--show-progress", action="store_true")
    parser.add_argument(
        "--routing-min-amount",
        type=float,
        default=1e-18,
        help=(
            "Minimum absolute temporal child amount to keep expanding during "
            "temporal_routing(). Larger values can prune tiny depth-6 branches."
        ),
    )
    parser.add_argument("--interpolation-start-year-offset", type=int, default=-20)
    parser.add_argument("--interpolation-end-year-offset", type=int, default=20)
    parser.add_argument("--no-cache-interpolation", action="store_true")
    parser.add_argument(
        "--solver-mode",
        choices=("bw2calc", "direct", "iterative"),
        default="iterative",
    )
    parser.add_argument(
        "--fallback-solver-mode",
        choices=("direct", "bw2calc", "none"),
        default="direct",
        help="Solver to retry with if iterative GMRES does not converge.",
    )
    parser.add_argument("--iterative-rtol", type=float, default=1e-3)
    parser.add_argument("--iterative-maxiter", type=int, default=300)
    parser.add_argument("--iterative-restart", type=int, default=50)
    return parser


if __name__ == "__main__":
    raise SystemExit(run(build_parser().parse_args()))
