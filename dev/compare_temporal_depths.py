from __future__ import annotations

import argparse
import csv
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
from datapackage import Package
from openpyxl import load_workbook
import plotly.graph_objects as go

from trails import Trails, plot_rf, plot_temp, plot_temporal_scores
from trails.fair_rf import run_fair_delta_rf

DEFAULT_METHODS = [
    "IPCC 2021 (incl. biogenic CO2) - climate change: total (incl. biogenic CO2) - global warming potential (GWP100)"
]


@dataclass(frozen=True)
class ActivityDef:
    name: str
    reference_product: str
    location: str


def _normalize(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _read_activity_and_exchanges(ws) -> tuple[ActivityDef | None, list[dict[str, str]]]:
    activity_name = ""
    reference_product = ""
    location = ""

    for r in range(1, 80):
        k = _normalize(ws.cell(r, 1).value).lower()
        v = _normalize(ws.cell(r, 2).value)
        if k == "activity":
            activity_name = v
        elif k == "reference product":
            reference_product = v
        elif k == "location":
            location = v

    if not activity_name:
        return None, []

    header_row = None
    for r in range(1, 120):
        if (
            _normalize(ws.cell(r, 1).value).lower() == "name"
            and _normalize(ws.cell(r, 10).value).lower() == "type"
        ):
            header_row = r
            break

    exchanges: list[dict[str, str]] = []
    if header_row is not None:
        headers = [_normalize(ws.cell(header_row, c).value) for c in range(1, 21)]
        for r in range(header_row + 1, 400):
            row_vals = [_normalize(ws.cell(r, c).value) for c in range(1, 21)]
            if not any(row_vals):
                continue
            row = {headers[i]: row_vals[i] for i in range(len(headers)) if headers[i]}
            exchanges.append(row)

    return ActivityDef(activity_name, reference_product, location), exchanges


def collect_terminal_activities(inventory_paths: Iterable[Path]) -> list[ActivityDef]:
    all_activities: list[ActivityDef] = []
    activity_names: set[str] = set()
    technosphere_supplier_names: list[str] = []

    for path in inventory_paths:
        wb = load_workbook(path, data_only=True)
        for ws in wb.worksheets:
            activity, exchanges = _read_activity_and_exchanges(ws)
            if activity is None:
                continue
            all_activities.append(activity)
            activity_names.add(activity.name)
            for exc in exchanges:
                if _normalize(exc.get("type", "")).lower() != "technosphere":
                    continue
                supplier_name = _normalize(exc.get("name", ""))
                if supplier_name:
                    technosphere_supplier_names.append(supplier_name)

    suppliers_used_by_imported = {
        supplier
        for supplier in technosphere_supplier_names
        if supplier in activity_names
    }

    terminal = [a for a in all_activities if a.name not in suppliers_used_by_imported]

    # Stable order, deduplicated
    seen: set[tuple[str, str, str]] = set()
    out: list[ActivityDef] = []
    for a in terminal:
        k = (a.name, a.reference_product, a.location)
        if k in seen:
            continue
        seen.add(k)
        out.append(a)
    return out


def _metadata_by_idx(trails: Trails) -> dict[int, dict]:
    if not trails.activity_indices:
        return {}
    first_label = next(iter(trails.activity_indices))
    by_idx = trails.activity_indices[first_label]
    return {int(k): v for k, v in by_idx.items()}


def match_activity_indices(
    trails: Trails, targets: list[ActivityDef]
) -> dict[ActivityDef, int]:
    by_idx = _metadata_by_idx(trails)
    out: dict[ActivityDef, int] = {}

    for target in targets:
        exact = []
        for idx, md in by_idx.items():
            if not isinstance(md, dict):
                continue
            if _normalize(md.get("name")) != target.name:
                continue
            if _normalize(md.get("reference product")) != target.reference_product:
                continue
            if _normalize(md.get("location")) != target.location:
                continue
            exact.append(idx)
        if len(exact) == 1:
            out[target] = exact[0]
            continue

        # Fallback to unique name match
        by_name = [
            idx
            for idx, md in by_idx.items()
            if isinstance(md, dict) and _normalize(md.get("name")) == target.name
        ]
        if len(by_name) == 1:
            out[target] = by_name[0]

    return out


def _to_method_vector(score_obj: object, methods: list[str]) -> np.ndarray:
    arr = np.asarray(score_obj, dtype=float)
    if arr.ndim == 0:
        return np.array([float(arr)], dtype=float)
    if arr.ndim == 1:
        return arr.astype(float, copy=False)
    return arr.ravel().astype(float, copy=False)


def _temporal_total_scores_by_method(trails: Trails, methods: list[str]) -> np.ndarray:
    if trails.scores is None:
        raise RuntimeError("trails.scores is None after temporal LCA run.")
    da = trails.scores
    if "method" in da.dims:
        reduce_dims = [d for d in da.dims if d != "method"]
        vals = da.sum(dim=reduce_dims).values
        return np.asarray(vals, dtype=float).reshape(-1)
    return np.array([float(da.sum().values)], dtype=float)


def _slugify(value: str) -> str:
    text = _normalize(value).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text or "item"


def _matches_substring(value: str, patterns: list[str]) -> bool:
    text = _normalize(value).lower()
    return any(p in text for p in patterns)


def _plot_static_score_payload(
    methods: list[str], static_vec: np.ndarray
) -> float | dict[str, float]:
    if len(methods) <= 1:
        return float(static_vec[0]) if static_vec.size else 0.0

    if static_vec.size == 0:
        return {method: 0.0 for method in methods}

    payload: dict[str, float] = {}
    for i, method in enumerate(methods):
        j = i if i < static_vec.size else static_vec.size - 1
        payload[method] = float(static_vec[j])
    return payload


def _export_depth_plots(
    *,
    trails: Trails,
    activity: ActivityDef,
    activity_index: int,
    depth: int,
    methods: list[str],
    static_vec: np.ndarray,
    reference_year: int,
    output_dir: Path,
) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)

    figs = plot_temporal_scores(
        trails=trails,
        stacked=False,
        legend_top_n=7,
        show_flow_contributions=False,
        title="",
        method_label="Impact score",
        cumulative=False,
        width=700,
        height=500,
        year_tick=5,
        year_range=(1990, 2100),
        reference_year=int(reference_year),
        show_cumulative_axis=True,
        static_score=_plot_static_score_payload(methods, static_vec),
        static_score_dash="dot",
        static_score_color="red",
    )

    fig_list = figs if isinstance(figs, list) else [figs]
    safe_name = _slugify(activity.name)
    paths: list[Path] = []
    for i, fig in enumerate(fig_list):
        method_name = methods[i] if i < len(methods) else f"method_{i + 1}"
        safe_method = _slugify(method_name)[:80]
        filename = f"idx_{int(activity_index)}_{safe_name}_depth_{int(depth)}_{safe_method}.png"
        out_path = output_dir / filename
        fig.write_image(str(out_path))
        paths.append(out_path)
    return paths


def _export_fair_plots(
    *,
    trails: Trails,
    activity: ActivityDef,
    activity_index: int,
    depth: int,
    reference_year: int,
    output_dir: Path,
) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_name = _slugify(activity.name)

    def _empty_fig(title: str, message: str) -> go.Figure:
        fig = go.Figure()
        fig.update_layout(
            width=700,
            height=500,
            title={"text": title},
            xaxis={"visible": False},
            yaxis={"visible": False},
        )
        fig.add_annotation(
            text=message,
            x=0.5,
            y=0.5,
            xref="paper",
            yref="paper",
            showarrow=False,
        )
        return fig

    try:
        rf_fig = plot_rf(
            trails=trails,
            by="flow",
            title="Radiative forcing by flow",
            method_label="W/m²",
            stacked=False,
            legend_top_n=7,
            width=700,
            height=500,
            year_tick=5,
            year_range=(1990, 2100),
            reference_year=int(reference_year),
            show_cumulative_axis=True,
            show_cumulative_quantile_band=False,
            flow_groupby_name=True,
        )
    except ValueError as exc:
        if "No scores_by_flow found." not in str(exc):
            raise
        try:
            rf_fig = plot_rf(
                trails=trails,
                by="root activity",
                title="Radiative forcing by root activity",
                method_label="W/m²",
                stacked=False,
                legend_top_n=7,
                width=700,
                height=500,
                year_tick=5,
                year_range=(1990, 2100),
                reference_year=int(reference_year),
                show_cumulative_axis=True,
                show_cumulative_quantile_band=False,
            )
        except ValueError as inner_exc:
            if "No scores_by_first_level_child found." not in str(inner_exc):
                raise
            rf_fig = _empty_fig(
                "Radiative forcing",
                "No allocatable radiative forcing contributions for this case.",
            )
    rf_path = (
        output_dir
        / f"idx_{int(activity_index)}_{safe_name}_depth_{int(depth)}_radiative_forcing.png"
    )
    rf_fig.write_image(str(rf_path))

    try:
        temp_fig = plot_temp(
            trails=trails,
            by="flow",
            title="Temperature anomaly by flow",
            method_label="°C",
            stacked=False,
            legend_top_n=7,
            width=700,
            height=500,
            year_tick=5,
            year_range=(1990, 2100),
            reference_year=int(reference_year),
            show_total_axis=True,
            show_total_quantile_band=False,
            flow_groupby_name=True,
        )
    except ValueError as exc:
        if "No scores_by_flow found." not in str(exc):
            raise
        try:
            temp_fig = plot_temp(
                trails=trails,
                by="root activity",
                title="Temperature anomaly by root activity",
                method_label="°C",
                stacked=False,
                legend_top_n=7,
                width=700,
                height=500,
                year_tick=5,
                year_range=(1990, 2100),
                reference_year=int(reference_year),
                show_total_axis=True,
                show_total_quantile_band=False,
            )
        except ValueError as inner_exc:
            if "No scores_by_first_level_child found." not in str(inner_exc):
                raise
            temp_fig = _empty_fig(
                "Temperature anomaly",
                "No allocatable temperature contributions for this case.",
            )
    temp_path = (
        output_dir
        / f"idx_{int(activity_index)}_{safe_name}_depth_{int(depth)}_temperature_anomaly.png"
    )
    temp_fig.write_image(str(temp_path))

    return [rf_path, temp_path]


def run(args: argparse.Namespace) -> int:
    package = Package(str(args.datapackage))
    trails = Trails(
        package,
        interpolate_annual=True,
        interpolation_start_year_offset=args.interpolation_start_year_offset,
        interpolation_end_year_offset=args.interpolation_end_year_offset,
        debug=False,
    )

    inventory_paths = [Path(p).resolve() for p in args.inventories]
    trails.import_excel_inventory([str(p) for p in inventory_paths])

    terminals = collect_terminal_activities(inventory_paths)
    excluded_substrings = [
        _normalize(v).lower() for v in args.exclude_activity_substring if _normalize(v)
    ]
    if excluded_substrings:
        terminals = [
            a for a in terminals if not _matches_substring(a.name, excluded_substrings)
        ]
    if not terminals:
        raise RuntimeError("No terminal imported activities left after filtering.")
    idx_map = match_activity_indices(trails, terminals)

    if not idx_map:
        raise RuntimeError(
            "No terminal imported activities could be matched to indices."
        )

    depths = sorted({int(d) for d in args.depths})
    if not depths:
        raise ValueError("At least one depth must be provided via --depths.")
    if any(d < 0 for d in depths):
        raise ValueError("Depth values must be non-negative integers.")
    methods = list(args.methods)
    plot_dir = Path(args.plot_dir).resolve()

    rows: list[dict[str, object]] = []
    total_acts = len(idx_map)
    for act_i, (act_def, idx) in enumerate(idx_map.items(), start=1):
        print(
            f"\n[{act_i}/{total_acts}] Activity idx={idx} | "
            f"{act_def.name} ({act_def.reference_product}, {act_def.location})"
        )

        t0_static = time.perf_counter()
        trails.static_lca(
            year=int(args.reference_year),
            act_idx=int(idx),
            methods=methods,
            amount=float(args.amount),
        )
        static_vec = _to_method_vector(trails.static_score, methods)
        dt_static = time.perf_counter() - t0_static
        print(f"  static_lca done in {dt_static:.2f}s")

        temporal_by_depth: dict[int, np.ndarray] = {}
        for depth in depths:
            print(f"  depth={depth}: temporal_routing + lca ...")
            t0_depth = time.perf_counter()
            trails.temporal_routing(
                start_year=int(args.reference_year),
                start_act_idx=int(idx),
                amount=float(args.amount),
                max_depth=int(depth),
                show_progress=bool(args.show_progress),
                attribute_to_roots=True,
            )
            trails.lca(
                methods=methods,
                show_progress=bool(args.show_progress),
                compute_score=True,
                store_inventory=bool(args.run_fair),
            )
            temporal_by_depth[depth] = _temporal_total_scores_by_method(trails, methods)
            plot_paths = _export_depth_plots(
                trails=trails,
                activity=act_def,
                activity_index=int(idx),
                depth=int(depth),
                methods=methods,
                static_vec=static_vec,
                reference_year=int(args.reference_year),
                output_dir=plot_dir,
            )
            print("  depth=" f"{depth}: wrote {len(plot_paths)} plot(s) to {plot_dir}")

            if bool(args.run_fair):
                print(f"  depth={depth}: run_fair_delta_rf + RF/temperature plots ...")
                t0_fair = time.perf_counter()
                run_fair_delta_rf(
                    trails,
                    scenario=str(args.fair_scenario),
                    per_species_runs=bool(args.fair_per_species_runs),
                    per_species_workers=(
                        int(args.fair_per_species_workers)
                        if args.fair_per_species_workers is not None
                        else None
                    ),
                    validate_emissions_delta=not bool(args.fair_skip_validation),
                )
                fair_paths = _export_fair_plots(
                    trails=trails,
                    activity=act_def,
                    activity_index=int(idx),
                    depth=int(depth),
                    reference_year=int(args.reference_year),
                    output_dir=plot_dir,
                )
                print(
                    "  depth="
                    f"{depth}: wrote {len(fair_paths)} FaIR plot(s) to {plot_dir}"
                )
                dt_fair = time.perf_counter() - t0_fair
                print(f"  depth={depth}: FaIR done in {dt_fair:.2f}s")

            dt_depth = time.perf_counter() - t0_depth
            print(f"  depth={depth}: done in {dt_depth:.2f}s")

        for m_i, method in enumerate(methods):
            static_score = float(static_vec[m_i if m_i < len(static_vec) else 0])
            row: dict[str, object] = {
                "activity_index": int(idx),
                "activity_name": act_def.name,
                "reference_product": act_def.reference_product,
                "location": act_def.location,
                "method": method,
                "static_score": static_score,
            }

            prev_depth: int | None = None
            prev_score: float | None = None
            for depth in depths:
                depth_arr = temporal_by_depth[depth]
                depth_score = float(depth_arr[m_i if m_i < len(depth_arr) else 0])
                row[f"temporal_depth_{depth}"] = depth_score
                row[f"delta_d{depth}_vs_static"] = depth_score - static_score
                if prev_depth is not None and prev_score is not None:
                    row[f"delta_d{depth}_vs_d{prev_depth}"] = depth_score - prev_score
                prev_depth = int(depth)
                prev_score = depth_score

            rows.append(row)

    out_path = Path(args.output).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0].keys())
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print("\nTerminal imported activities analyzed:")
    for act_def, idx in sorted(idx_map.items(), key=lambda x: x[1]):
        print(
            f"- idx={idx} | {act_def.name} ({act_def.reference_product}, {act_def.location})"
        )

    print(f"\nWrote results: {out_path}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Compare static LCA and temporal LCA scores at routing depths 1 and 5 "
            "for terminal imported activities."
        )
    )
    parser.add_argument(
        "--datapackage",
        type=Path,
        default=Path(
            "/Users/romain/Library/CloudStorage/OneDrive-PaulScherrerInstitut/trails/data/trails_2026-03-03.zip"
        ),
        help="Path to Frictionless datapackage zip/json.",
    )
    parser.add_argument(
        "--inventories",
        nargs="+",
        default=[
            str(
                Path(
                    "/Users/romain/GitHub/trails/dev/lci-case-study-daccs_storage_risk.xlsx"
                )
            ),
            str(
                Path(
                    "/Users/romain/GitHub/trails/dev/lci-case-study-fertilizer_n2o_timing.xlsx"
                )
            ),
            str(
                Path(
                    "/Users/romain/GitHub/trails/dev/lci-case-study-marine_fuel_switch.xlsx"
                )
            ),
            str(
                Path(
                    "/Users/romain/GitHub/trails/dev/lci-case-study-biomass_growth_vs_gas_heat.xlsx"
                )
            ),
            str(
                Path(
                    "/Users/romain/GitHub/trails/dev/lci-case-study-ccu_polyol_delayed_release.xlsx"
                )
            ),
        ],
        help="Excel inventory files to import.",
    )
    parser.add_argument(
        "--methods",
        nargs="+",
        default=DEFAULT_METHODS,
        help="LCIA method names.",
    )
    parser.add_argument("--reference-year", type=int, default=2035)
    parser.add_argument("--amount", type=float, default=1.0)
    parser.add_argument(
        "--depths",
        nargs="+",
        type=int,
        default=[1, 5],
        help="Temporal routing depths to evaluate (e.g., --depths 5).",
    )
    parser.add_argument("--show-progress", action="store_true")
    parser.add_argument(
        "--exclude-activity-substring",
        nargs="+",
        default=["natural gas boiler"],
        help=(
            "Skip terminal activities whose activity name contains any of these "
            "case-insensitive substrings."
        ),
    )
    parser.add_argument("--interpolation-start-year-offset", type=int, default=-20)
    parser.add_argument("--interpolation-end-year-offset", type=int, default=20)
    parser.add_argument(
        "--run-fair",
        dest="run_fair",
        action="store_true",
        help="Run FaIR and export RF/temperature anomaly plots for each activity/depth.",
    )
    parser.add_argument(
        "--skip-fair",
        dest="run_fair",
        action="store_false",
        help="Skip FaIR run and only export temporal score plots.",
    )
    parser.set_defaults(run_fair=True)
    parser.add_argument(
        "--fair-scenario",
        type=str,
        default="REMIND|SSP2-PkBudg650",
        help="Scenario name present in FaIR emissions input.",
    )
    parser.add_argument(
        "--fair-per-species-runs",
        action=argparse.BooleanOptionalAction,
        default=False,
        help=(
            "Use per-species FaIR perturbation runs (slower, more exact "
            "decomposition)."
        ),
    )
    parser.add_argument(
        "--fair-per-species-workers",
        type=int,
        default=None,
        help="Optional worker count for per-species FaIR runs.",
    )
    parser.add_argument(
        "--fair-skip-validation",
        action="store_true",
        help="Skip FaIR emissions delta validation checks.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("/Users/romain/GitHub/trails/dev/temporal_depth_comparison.csv"),
        help="Output CSV path.",
    )
    parser.add_argument(
        "--plot-dir",
        type=Path,
        default=Path("/Users/romain/GitHub/trails/dev/temporal_depth_plots"),
        help="Directory for per-depth PNG plots.",
    )
    return parser


if __name__ == "__main__":
    parser = build_parser()
    raise SystemExit(run(parser.parse_args()))
